# Версия 2.0 модуля чтения данных из файлов Контингента и ActiveDirectory.
# Изменена структура данных:
# Вместо словарей со списком учеников используется список со словарями. Так типа удобней.

from openpyxl import load_workbook
import csv
from kz2eng import translate_that_shit, convert_name

#  Статусы учащихся
STILL_STUDYING = "Учится"
EXPORTED_TO_NIS = "Выбыл по переводу в НИШ"
GONE1 = "Выбыл по собственному желанию"
GONE2 = "Отчислен"
EXTERN = "На экстернате"

# Читаем данные из Контингента
def read_contingent():
    # Список для чтения из Контингента
    Contingent_List = []

    xl = load_workbook("Контингент.xlsx")
    sheet = xl.active
    for i in range(5, sheet.max_row + 1):
        # Словарь для чтения из Контингента
        Contingent_Dict = {
            'SamAccountName': '',
            'GivenName': '',
            'Surname': '',
            'Grade': '',
            'Status': ''
        }
        Name = str(sheet.cell(i, 3).value).split(" ")
        Surname = Name[0]
        GivenName = Name[1]
        Grade = str(sheet.cell(i, 7).value)
        if Grade == "None":
            Grade = ""
        Contingent_Dict['SamAccountName'] = str(sheet.cell(i, 4).value)
        Contingent_Dict['GivenName'] = GivenName
        Contingent_Dict['Surname'] = Surname
        Contingent_Dict['Grade'] = Grade
        Contingent_Dict['Status'] = str(sheet.cell(i, 10).value)
        Contingent_List.append(Contingent_Dict)
    keys = []
    for k in Contingent_Dict.keys():
        keys.append(k)
#    Contingent_List.insert(0,keys)
    return Contingent_List

# Читаем данные из ActiveDirectory
def read_AD():
    # Список для чтения из Контингента
    AD_List = []

    with open('exported.csv', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['SamAccountName', 'GivenName', 'Surname', 'Description', 'UserPrincipalName']
        ad = csv.DictReader(csvfile, fieldnames=fieldnames, delimiter=";")
        keys = []
        for key in fieldnames:
            keys.append(key)
        for item in ad:
            # Словарь для чтения из ActiveDirectory
            AD_Dict = {
                'SamAccountName': '',
                'GivenName': '',
                'Surname': '',
                'Description': '',
                'UserPrincipalName': ''
            }
            for k in keys:
                AD_Dict[k] = item[k]
            AD_List.append(AD_Dict)
    return AD_List

# Функция заполнения данных в Excel
def fill_data(data, starty, startx, sheet_out, header=True):

    if header == True:
        m = startx
        if len(data)>0:
            for k in data[0].keys():
                sheet_out.cell(starty, m, k)
                m += 1
        starty += 1

    m = starty
    for i in data:
        n = startx
        for d in i.keys():
            sheet_out.cell(m, n, i[d])
            n += 1
        m += 1

# Функция заполнения данных в csv
def write_to_csv(data, filename, header=True):
    if filename[-4:] != '.csv':
        filename += '.csv'
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        keys = []
        if len(data)>0:
            for k in data[0].keys():
                keys.append(k)
        c = csv.writer(csvfile, delimiter=";")
        if header:
            c.writerow(keys)
        for item in data:
            row = []
            if type(item)!=list:
                for key in keys:
                   row.append(item[key])
            c.writerow(row)

# Функция заполнения данных в csv с переводом названии учеток
def write_to_csv_translated(data, filename, givenname, surname, iin):
    if filename[-4:] != '.csv':
        filename += '.csv'
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        keys = []
        if len(data)>0:
            for k in data[0].keys():
                keys.append(k)
        keys.append("Translated")
        c = csv.writer(csvfile, delimiter=";")
        c.writerow(keys)
        for item in data:
            row = []
            if type(item) != list:
                item["Translated"] = translate_that_shit(item[surname]) + "_" + convert_name(item[givenname]) + item[iin][:4]
                for key in keys:
                    row.append(item[key])
            c.writerow(row)

def get_data_keys_count(data):
    if len(data)>0:
        keys = 0
        for k in data[0].keys():
            keys += 1
        return keys
    else:
        return 0