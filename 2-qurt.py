# Я люблю құрт.
# Иногда.
import read_data
from read_data import fill_data, write_to_csv, write_to_csv_translated, get_data_keys_count
from read_data import read_contingent, read_AD
from openpyxl import Workbook

# Словарь переведенных между классами
ChangedClass_Dict = {
    'SamAccountName': '',
    'GivenName': '',
    'Surname': '',
    'LastGrade': '',
    'NewGrade' : '',
    'UserPrincipalName': ''
}

# Список для отсутствующих в ActiveDirectory учеников
Absent_List = []

# Список для выбывших учеников
Exported_List = []

# Список переведенных между классами
ChangedClass_List = []

# Читаем данные из Контингента
cont = read_contingent()

# Читаем данные из ActiveDirectory
ad = read_AD()
keys = []

# Заполняем список отсутствующих в ActiveDirectory учеников
for c in cont:
    exists = False
    for a in ad:
        if c["SamAccountName"] == a["SamAccountName"]:
            exists = True
            break
    if (exists == False) and (c["Status"] == read_data.STILL_STUDYING):
        Absent_List.append(c)

# Заполняем список выбывших по Контингенту учеников
for c in cont:
    for a in ad:
        if (c["SamAccountName"] == a["SamAccountName"]) and ((c["Status"] == read_data.EXPORTED_TO_NIS) or (c["Status"] == read_data.GONE1) or (c["Status"] == read_data.GONE2)):
            Exported_List.append(c)
            break

# Заполняем список переведенных между классами
keys = []
for key in ChangedClass_Dict.keys():
    keys.append(key)
for c in cont:
    for a in ad:
        if (c["SamAccountName"] == a["SamAccountName"]):
            if (c["Grade"] != a["Description"]) and (c["Status"] == read_data.STILL_STUDYING):
                for k in keys:
                    ChangedClass_Dict[k] = ''
                ChangedClass_Dict["SamAccountName"] = c["SamAccountName"]
                ChangedClass_Dict["GivenName"] = c["GivenName"]
                ChangedClass_Dict["Surname"] = c["Surname"]
                ChangedClass_Dict["LastGrade"] = a["Description"]
                ChangedClass_Dict["NewGrade"] = c["Grade"]
                ChangedClass_Dict["UserPrincipalName"] = a["UserPrincipalName"]
                ChangedClass_List.append(ChangedClass_Dict)
            break

# Заполняем таблицу в Excel
print("Учеников, отсутствующих в ActiveDirectory:", len(Absent_List))
print("Учеников, выбывших по Контингенту:", len(Exported_List))
print("Учеников, переведенных между классами:", len(ChangedClass_List))
xl_out = Workbook()
sheet = xl_out.active
x = -4
if len(Absent_List)>0:
    x += get_data_keys_count(Absent_List)
    sheet.cell(1, x, "Отсутствующие в ActiveDirectory ученики")
    fill_data(Absent_List, 2, x, sheet)
    filename = input("Название файла с отсутствующими в ActiveDirectory:")
    write_to_csv_translated(Absent_List, filename, "GivenName", "Surname", "SamAccountName")
if len(Exported_List)>0:
    x += get_data_keys_count(Exported_List)
    sheet.cell(1, x, "Выбывшие по Контингенту ученики")
    fill_data(Exported_List, 2, x, sheet)
    filename = input("Название файла с выбывшими по Контингенту:")
    write_to_csv(Exported_List, filename)
if len(ChangedClass_List)>0:
    x += get_data_keys_count(ChangedClass_List)
    sheet.cell(1, x, "Переведенные между классами")
    fill_data(ChangedClass_List, 2, x, sheet)
    filename = input("Название файла с переведенными между классами:")
    write_to_csv(ChangedClass_List, filename)

# Сохраняем в общий Excel-файл
xl_out.save("STUDENTS_DETAILED.xlsx")
xl_out.close()
print("Таблица с изменениями в ActiveDirectory и Контингенте сохранены в файле STUDENTS_DETAILED.xlsx")
# Сохраняем все это счастье в CSV-файлы
print("Ешь курт!")
u = input("Нажми Enter, и вали отсюда")
